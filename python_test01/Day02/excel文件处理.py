"""
Python 操作 Excel 需要三方库的支持，如果要兼容 Excel 2007 以前的版本，也就是xls格式的 Excel 文件，可以使用三方库xlrd和xlwt，前者用于读 Excel 文件，后者用于写 Excel 文件。
如果使用较新版本的 Excel，即xlsx格式的 Excel 文件，可以使用openpyxl库，当然这个库不仅仅可以操作Excel，还可以操作其他基于 Office Open XML 的电子表格文件。

前面基于xlwt和xlrd操作 Excel 文件，大家可以先使用下面的命令安装这两个三方库以及配合使用的工具模块xlutils。
pip install xlwt xlrd xlutils
"""
import random
import datetime

import openpyxl
import xlrd
import xlwt
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook import Workbook
from xlutils.copy import copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


def excel_create():
    """
    流程：
    1. 创建工作簿对象 xlwt.Workbook()
    2. 创建工作表  wb.add_sheet('学生成绩表')
    3. 设置样式
    4. 写入内容 ws.write(行, 列, 内容, 样式)
    5. 保存 wb.save('../../data/raw/students.xls')

    样式：
    字体（Font）、
    对齐方式（Alignment）、
    边框（Border）
    背景（Background）
    xlwt 通过XFStyle 进行设置
    :return:
    """
    student_names = ['关羽', '张飞', '赵云', '马超', '黄忠']
    scores = [[random.randrange(50, 101) for _ in range(3)] for _ in range(5)]
    # 创建工作簿对象（Workbook）
    wb = xlwt.Workbook()
    # 创建工作表对象（Worksheet）
    ws = wb.add_sheet('学生成绩表')
    # 表头样式
    header_style = xlwt.XFStyle()

    #背景
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    # 0 - 黑色、1 - 白色、2 - 红色、3 - 绿色、4 - 蓝色、5 - 黄色、6 - 粉色、7 - 青色
    pattern.pattern_fore_colour = 5

    #字体
    font = xlwt.Font()
    # 字体名称
    font.name = '华文楷体'
    # 字体大小（20是基准单位，18表示18px）
    font.height = 20 * 18
    # 是否使用粗体
    font.bold = True
    # 是否使用斜体
    font.italic = False
    # 字体颜色
    font.colour_index = 0

    #对齐
    align = xlwt.Alignment()
    # 垂直方向的对齐方式
    align.vert = xlwt.Alignment.VERT_CENTER
    # 水平方向的对齐方式
    align.horz = xlwt.Alignment.HORZ_CENTER

    #边框
    borders = xlwt.Borders()
    props = (
        ('top', 'top_colour'), ('right', 'right_colour'),
        ('bottom', 'bottom_colour'), ('left', 'left_colour')
    )
    # 通过循环对四个方向的边框样式及颜色进行设定
    for position, color in props:
        # 使用setattr内置函数动态给对象指定的属性赋值
        setattr(borders, position, xlwt.Borders.DASHED)
        setattr(borders, color, 5)

    header_style.borders = borders
    header_style.alignment = align
    header_style.font = font
    header_style.pattern = pattern
    # 添加表头数据
    # 设置行宽
    # 设置行高为40px
    ws.row(0).set_style(xlwt.easyxf(f'font:height {20 * 40}'))
    titles = ('姓名', '语文', '数学', '英语')
    for index, title in enumerate(titles):
        # 设置列宽为200px
        ws.col(index).width = 20 * 200
        # 设置单元格的数据和样式
        ws.write(0, index, title, header_style)
    # 将学生姓名和考试成绩写入单元格
    for row in range(len(scores)):
        ws.write(row + 1, 0, student_names[row])
        for col in range(len(scores[row])):
            ws.write(row + 1, col + 1, scores[row][col])
    # 保存工作簿对象
    wb.save('../../data/raw/students.xls')
    pass

def excel_read():
    """

    :return:
    """
    # 使用xlrd模块的open_workbook函数打开指定Excel文件并获得Book对象（工作簿）
    wb = xlrd.open_workbook('../../data/raw/students.xls')
    # 通过Book对象的sheet_names方法可以获取所有表单名称
    sheetnames = wb.sheet_names()
    print(sheetnames)
    # 通过指定的表单名称获取Sheet对象（工作表）
    sheet = wb.sheet_by_name(sheetnames[0])
    # 通过Sheet对象的nrows和ncols属性获取表单的行数和列数
    print(sheet.nrows, sheet.ncols)
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            # 通过Sheet对象的cell方法获取指定Cell对象（单元格）
            # 通过Cell对象的value属性获取单元格中的值
            value = sheet.cell(row, col).value
            # 对除首行外的其他行进行数据格式化处理
            # if row > 0:
            #     # 第1列的xldate类型先转成元组再格式化为“年月日”的格式
            #     if col == 0:
            #         # xldate_as_tuple函数的第二个参数只有0和1两个取值
            #         # 其中0代表以1900-01-01为基准的日期，1代表以1904-01-01为基准的日期
            #         value = xlrd.xldate_as_tuple(value, 0)
            #         value = f'{value[0]}年{value[1]:>02d}月{value[2]:>02d}日'
            #     # 其他列的number类型处理成小数点后保留两位有效数字的浮点数
            #     else:
            #         value = f'{value:.2f}'
            print(value, end='\t')
        print()
    # 获取最后一个单元格的数据类型
    # 0 - 空值，1 - 字符串，2 - 数字，3 - 日期，4 - 布尔，5 - 错误
    last_cell_type = sheet.cell_type(sheet.nrows - 1, sheet.ncols - 1)
    print(last_cell_type)
    # 获取第一行的值（列表）
    print(sheet.row_values(0))
    # 获取指定行指定列范围的数据（列表）
    # 第一个参数代表行索引，第二个和第三个参数代表列的开始（含）和结束（不含）索引
    print(sheet.row_slice(3, 0, 2))

def formula():
    """
    1. 打开文件，获取sheet,获取行列
    2. 复制文件，在新的文件里写入，（行，列，格式化后的内容）
    :return:
    """
    wb = xlrd.open_workbook('../../data/raw/students.xls')
    sheet = wb.sheet_by_index(0)
    nrows, ncols = sheet.nrows, sheet.ncols
    wb_for_write = copy(wb)
    sheet2 = wb_for_write.get_sheet(0)
    sheet2.write(nrows, 4, xlwt.Formula(f'average(B2:E{nrows})'))
    # sheet2.write(nrows, 6, xlwt.Formula(f'sum(G2:G{nrows})'))
    wb_for_write.save('../../data/raw/students汇总.xls')

"""
openpyxl
"""
def xlsx_create():
    # 第一步：创建工作簿（Workbook）
    wb = openpyxl.Workbook()
    # 第二步：添加工作表（Worksheet）
    sheet = wb.active
    sheet.title = '期末成绩'
    titles = ('姓名', '语文', '数学', '英语')
    for col_index, title in enumerate(titles):
        sheet.cell(1, col_index + 1, title)
    names = ('关羽', '张飞', '赵云', '马超', '黄忠')
    for row_index, name in enumerate(names):
        sheet.cell(row_index + 2, 1, name)
        for col_index in range(2, 5):
            sheet.cell(row_index + 2, col_index, random.randrange(50, 101))
    # 第四步：保存工作簿
    wb.save('../../data/raw/考试成绩表.xlsx')

def xlsx_read():
    # 加载一个工作簿 ---> Workbook
    wb = openpyxl.load_workbook('../../data/raw/考试成绩表.xlsx')
    # 获取工作表的名字
    print(wb.sheetnames)
    # 获取工作表 ---> Worksheet
    sheet = wb.worksheets[0]
    # 获得单元格的范围
    print(sheet.dimensions)
    # 获得行数和列数
    print(sheet.max_row, sheet.max_column)
    # 获取指定单元格的值
    print(sheet.cell(3, 3).value)
    print(sheet['C3'].value)
    print(sheet['G255'].value)
    # 读取所有单元格的数据
    for row_ch in range(2, sheet.max_row + 1):
        for col_ch in 'ABCDEFG':
            value = sheet[f'{col_ch}{row_ch}'].value
            if type(value) == datetime.datetime:
                print(value.strftime('%Y年%m月%d日'), end='\t')
            elif type(value) == int:
                print(f'{value:<10d}', end='\t')
            elif type(value) == float:
                print(f'{value:.4f}', end='\t')
            else:
                print(value, end='\t')
        print()

def excel_foramt():
    # 对齐方式
    alignment = Alignment(horizontal='center', vertical='center')
    # 边框线条
    side = Side(color='ff7f50', style='mediumDashed')

    wb = openpyxl.load_workbook('../../data/raw/考试成绩表.xlsx')
    sheet = wb.worksheets[0]

    # 调整行高和列宽
    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['E'].width = 120

    sheet['E1'] = '平均分'
    # 设置字体
    sheet.cell(1, 5).font = Font(size=18, bold=True, color='ff1493', name='华文楷体')
    # 设置对齐方式
    sheet.cell(1, 5).alignment = alignment
    # 设置单元格边框
    sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
    for i in range(2, 7):
        # 公式计算每个学生的平均分
        sheet[f'E{i}'] = f'=average(B{i}:D{i})'
        sheet.cell(i, 5).font = Font(size=12, color='4169e1', italic=True)
        sheet.cell(i, 5).alignment = alignment

    wb.save('../../data/raw/考试成绩表.xlsx')

def insert_chart():
    wb = Workbook(write_only=True)
    sheet = wb.create_sheet()

    rows = [
        ('类别', '销售A组', '销售B组'),
        ('手机', 40, 30),
        ('平板', 50, 60),
        ('笔记本', 80, 70),
        ('外围设备', 20, 10),
    ]

    # 向表单中添加行
    for row in rows:
        sheet.append(row)

    # 创建图表对象
    chart = BarChart()
    chart.type = 'col'
    chart.style = 10
    # 设置图表的标题
    chart.title = '销售统计图'
    # 设置图表纵轴的标题
    chart.y_axis.title = '销量'
    # 设置图表横轴的标题
    chart.x_axis.title = '商品类别'
    # 设置数据的范围
    data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)
    # 设置分类的范围
    cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
    # 给图表添加数据
    chart.add_data(data, titles_from_data=True)
    # 给图表设置分类
    chart.set_categories(cats)
    chart.shape = 4
    # 将图表添加到表单指定的单元格中
    sheet.add_chart(chart, 'A10')

    wb.save('../../data/raw/demo.xlsx')

if __name__ == '__main__':
    insert_chart()
    # for i in range(2,4):
    #     print(i)